"""
PDF 转换服务

支持将 Word(docx)、Excel(xlsx)、Txt、Markdown(md) 格式转换为 PDF
策略：所有格式先转为 Markdown，再通过 Markdown 转 PDF
"""
import io
import logging
import platform
from pathlib import Path
from typing import List, Tuple

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)


class PDFConverterService:
    """PDF 转换服务"""

    def __init__(self):
        self.supported_formats = ["docx", "xlsx", "txt", "md"]
        self._font_name = None
        self._styles = None
        self._page_width = None
        self._page_height = None
        self._setup_fonts()

    def _setup_fonts(self):
        """设置字体"""
        try:
            self._page_width, self._page_height = A4

            # 查找中文字体
            font_path = self._find_chinese_font()
            if font_path:
                try:
                    font = TTFont('ChineseFont', font_path)
                    pdfmetrics.registerFont(font)
                    from reportlab.pdfbase.pdfmetrics import registerFontFamily
                    registerFontFamily('ChineseFont', normal='ChineseFont')
                    self._font_name = 'ChineseFont'
                    logger.info(f"成功注册中文字体: {font_path}")
                except Exception as e:
                    logger.warning(f"字体注册失败: {e}, 使用Helvetica")
                    self._font_name = 'Helvetica'
            else:
                self._font_name = 'Helvetica'
                logger.warning("未找到中文字体，使用 Helvetica（不支持中文）")

            # 创建样式
            styles = getSampleStyleSheet()

            styles.add(ParagraphStyle(
                name='ChineseTitle',
                fontName=self._font_name,
                fontSize=16,
                leading=22,
                alignment=TA_CENTER,
                spaceAfter=12,
            ))

            styles.add(ParagraphStyle(
                name='ChineseHeading',
                fontName=self._font_name,
                fontSize=14,
                leading=20,
                spaceBefore=10,
                spaceAfter=8,
            ))

            styles.add(ParagraphStyle(
                name='ChineseBody',
                fontName=self._font_name,
                fontSize=10,
                leading=14,
                alignment=TA_JUSTIFY,
                spaceAfter=6,
            ))

            styles.add(ParagraphStyle(
                name='ChineseCode',
                fontName='Courier',
                fontSize=9,
                leading=12,
            ))

            self._styles = styles
            logger.info("PDF服务初始化完成")

        except Exception as e:
            logger.error(f"PDF服务初始化失败: {e}")
            raise

    def _find_chinese_font(self) -> str:
        """查找中文字体"""
        system = platform.system()

        if system == "Windows":
            fonts = [
                "C:/Windows/Fonts/simhei.ttf",
                "C:/Windows/Fonts/simsun.ttc",
                "C:/Windows/Fonts/msyh.ttc",
                "C:/Windows/Fonts/simsun.ttf",
            ]
        elif system == "Darwin":
            fonts = [
                "/System/Library/Fonts/STHeiti Light.ttc",
                "/System/Library/Fonts/PingFang.ttc",
                "/Library/Fonts/Arial Unicode.ttf",
            ]
        else:
            fonts = [
                "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            ]

        for font in fonts:
            if Path(font).exists():
                return font
        return None

    def _sanitize_text(self, text: str) -> str:
        """清理文本"""
        if not text:
            return ""
        return text.replace('\x00', '')

    async def convert_to_pdf(
        self,
        file_content: bytes,
        source_format: str,
        filename: str = "document"
    ) -> Tuple[bytes, str]:
        """将文档转换为 PDF"""
        try:
            if source_format.lower() not in self.supported_formats:
                return b"", f"不支持的格式: {source_format}"

            # 第一步：转换为 Markdown
            markdown_content, error = await self._convert_to_markdown(file_content, source_format, filename)
            if error:
                return b"", error

            # 第二步：Markdown 转 PDF
            return await self._convert_markdown_to_pdf(markdown_content, filename)

        except Exception as e:
            logger.error(f"PDF转换失败: {e}")
            import traceback
            logger.error(f"详细错误: {traceback.format_exc()}")
            return b"", f"转换失败: {str(e)}"

    async def _convert_to_markdown(
        self,
        file_content: bytes,
        source_format: str,
        filename: str
    ) -> Tuple[str, str]:
        """将各种格式转换为 Markdown"""
        converters = {
            "docx": self._convert_docx_to_markdown,
            "xlsx": self._convert_xlsx_to_markdown,
            "txt": self._convert_txt_to_markdown,
            "md": self._convert_md_to_markdown,
        }
        return await converters[source_format.lower()](file_content, filename)

    async def _convert_txt_to_markdown(self, file_content: bytes, filename: str) -> Tuple[str, str]:
        """Txt 转 Markdown"""
        try:
            text = self._decode_content(file_content)
            text = self._sanitize_text(text)
            return f"# {filename}\n\n{text}", ""
        except Exception as e:
            logger.error(f"Txt转Markdown失败: {e}")
            return "", f"文本文件处理失败: {str(e)}"

    async def _convert_md_to_markdown(self, file_content: bytes, filename: str) -> Tuple[str, str]:
        """Markdown 原样返回"""
        try:
            content = self._decode_content(file_content)
            content = self._sanitize_text(content)
            return f"# {filename}\n\n{content}", ""
        except Exception as e:
            logger.error(f"Markdown处理失败: {e}")
            return "", f"Markdown处理失败: {str(e)}"

    async def _convert_docx_to_markdown(self, file_content: bytes, filename: str) -> Tuple[str, str]:
        """Word 转 Markdown - 使用 zipfile 直接解析，更加健壮"""
        try:
            import zipfile
            import re

            lines = [f"# {filename}", ""]

            # 直接使用 zipfile 解析 DOCX，避免 python-docx 的严格验证
            try:
                with zipfile.ZipFile(io.BytesIO(file_content), 'r') as zf:
                    # 读取主文档内容
                    xml_content = zf.read('word/document.xml').decode('utf-8')
            except zipfile.BadZipFile:
                return "", "文件不是有效的 DOCX 格式"
            except KeyError:
                return "", "DOCX 文件损坏：找不到 document.xml"

            # 简单的 XML 解析 - 提取文本段落
            # 移除 XML 标签，提取纯文本
            xml_content = re.sub(r'<w:br[^>]*>', '\n', xml_content)
            xml_content = re.sub(r'</w:p>', '\n', xml_content)
            xml_content = re.sub(r'<[^>]+>', '', xml_content)
            xml_content = re.sub(r'\n\s*\n', '\n\n', xml_content)

            # 解码 HTML 实体
            xml_content = xml_content.replace('&amp;', '&')
            xml_content = xml_content.replace('&lt;', '<')
            xml_content = xml_content.replace('&gt;', '>')
            xml_content = xml_content.replace('&quot;', '"')
            xml_content = xml_content.replace('&#39;', "'")

            # 清理空白
            lines_text = [line.strip() for line in xml_content.split('\n') if line.strip()]

            # 生成 Markdown
            for text in lines_text[:500]:  # 限制最多500行
                if text:
                    lines.append(text)

            return '\n'.join(lines), ""

        except Exception as e:
            logger.error(f"Word转Markdown失败: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return "", f"Word文档处理失败: {str(e)}"
            for table in doc.tables:
                lines.append("")
                for row in table.rows:
                    row_data = [cell.text.strip() for cell in row.cells]
                    lines.append("| " + " | ".join(row_data) + " |")
                # 表头分隔符
                if table.rows:
                    lines.append("| " + " | ".join(["---"] * len(table.rows[0].cells)) + " |")

            return "\n".join(lines), ""

        except Exception as e:
            logger.error(f"Word转Markdown失败: {e}")
            return "", f"Word文档处理失败: {str(e)}"

    async def _convert_xlsx_to_markdown(self, file_content: bytes, filename: str) -> Tuple[str, str]:
        """Excel 转 Markdown"""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            lines = [f"# {filename} - Excel数据", ""]

            for sheet_name in wb.sheetnames[:10]:
                ws = wb[sheet_name]
                lines.append(f"## 工作表: {sheet_name}")
                lines.append("")

                for row_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if not any(row_data):
                        continue
                    lines.append("| " + " | ".join(row_data) + " |")
                    if row_idx == 0:
                        lines.append("| " + " | ".join(["---"] * len(row_data)) + " |")

                lines.append("")

            return "\n".join(lines), ""

        except Exception as e:
            logger.error(f"Excel转Markdown失败: {e}")
            return "", f"Excel处理失败: {str(e)}"

    async def _convert_markdown_to_pdf(self, markdown_content: str, filename: str) -> Tuple[bytes, str]:
        """Markdown 转 PDF"""
        try:
            logger.info(f"Markdown转PDF开始 - filename={filename}, 字体={self._font_name}")
            logger.info(f"styles['ChineseTitle'].fontName={self._styles['ChineseTitle'].fontName}")

            buffer = io.BytesIO()
            story = []

            safe_filename = self._sanitize_text(filename)
            logger.info(f"safe_filename={repr(safe_filename[:50])}")

            story.append(Paragraph(text=safe_filename, style=self._styles['ChineseTitle']))
            story.append(Spacer(1, 12))

            in_code = False
            for line in markdown_content.split('\n'):
                line = line.strip()

                if line.startswith('```'):
                    in_code = not in_code
                    story.append(Spacer(1, 6))
                    continue

                if in_code:
                    story.append(Paragraph(text=self._sanitize_text(line), style=self._styles['ChineseCode']))
                    continue

                if not line:
                    story.append(Spacer(1, 6))
                    continue

                # 标题处理
                if line.startswith('# '):
                    story.append(Paragraph(text=self._sanitize_text(line[2:]), style=self._styles['ChineseHeading']))
                elif line.startswith('## '):
                    story.append(Paragraph(text=self._sanitize_text(line[3:]), style=self._styles['ChineseHeading']))
                elif line.startswith('### '):
                    story.append(Paragraph(text=self._sanitize_text(line[4:]), style=self._styles['ChineseHeading']))
                elif line.startswith('#### '):
                    story.append(Paragraph(text=self._sanitize_text(line[5:]), style=self._styles['ChineseHeading']))
                elif line.startswith('- ') or line.startswith('* '):
                    story.append(Paragraph(text="• " + self._sanitize_text(line[2:]), style=self._styles['ChineseBody']))
                # 表格处理
                elif line.startswith('|'):
                    # 跳过 markdown 表格分隔符
                    if set(line.replace('|', '').replace('-', '').replace(':', '').replace(' ', '')) == set():
                        continue
                    # 解析并创建表格
                    table_lines = []
                    for _ in range(50):  # 最多50行
                        if line.startswith('|'):
                            row = [cell.strip() for cell in line.split('|')[1:-1]]
                            if not any(row) or set(''.join(row).replace('-', '').replace(':', '').replace(' ', '')) == set():
                                break
                            table_lines.append(row)
                            try:
                                line = next(markdown_content.split('\n').__iter__()).strip()
                            except StopIteration:
                                break
                        else:
                            break

                    if table_lines:
                        # 创建表格
                        t = Table(table_lines, colWidths=[100] * len(table_lines[0]))
                        t.setStyle(TableStyle([
                            ('FONTNAME', (0, 0), (-1, -1), self._font_name),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, '#999999'),
                            ('BACKGROUND', (0, 0), (-1, 0), '#4472C4'),
                            ('TEXTCOLOR', (0, 0), (-1, 0), '#FFFFFF'),
                        ]))
                        story.append(t)
                        story.append(Spacer(1, 6))
                else:
                    story.append(Paragraph(text=self._sanitize_text(line), style=self._styles['ChineseBody']))

            logger.info(f"准备构建PDF，story长度={len(story)}")

            pdf_doc = SimpleDocTemplate(
                buffer,
                pagesize=(self._page_width, self._page_height),
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            logger.info("调用pdf_doc.build()")
            pdf_doc.build(story)
            logger.info("pdf_doc.build()完成")

            result = buffer.getvalue()
            buffer.close()
            return result, ""

        except Exception as e:
            logger.error(f"Markdown转PDF失败: {e}")
            import traceback
            logger.error(f"详细错误: {traceback.format_exc()}")
            return b"", f"Markdown转PDF失败: {str(e)}"

    def _decode_content(self, file_content: bytes) -> str:
        """解码文件内容"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin-1']
        for enc in encodings:
            try:
                return file_content.decode(enc)
            except (UnicodeDecodeError, LookupError):
                continue
        return file_content.decode('utf-8', errors='replace')

    def get_supported_formats(self) -> List[str]:
        """获取支持的格式"""
        return self.supported_formats


# 全局单例
pdf_converter_service = PDFConverterService()